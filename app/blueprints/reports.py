from datetime import date, datetime, timedelta
from flask import Blueprint, render_template, request, send_file, flash
from flask_login import login_required
from io import StringIO, BytesIO
import csv

from app.extensions import db
from app.models.core import Site

bp = Blueprint("reports", __name__)

def _parse_date(s: str | None, default: date) -> date:
    if not s: return default
    y,m,d = map(int, s.split("-"))
    return date(y,m,d)


@bp.route("/", methods=["GET", "POST"])
@login_required
def index():
    sites = Site.query.order_by(Site.name).all()
    # default: prethodnih 7 dana (bez današnjeg u toku)
    today = datetime.utcnow().date()
    d_to = _parse_date(request.values.get("to"), today - timedelta(days=1))
    d_from = _parse_date(request.values.get("from"), d_to - timedelta(days=6))
    site_id = request.values.get("site_id", type=int)

    rows = []
    kpis = {"today": 0.0, "mtd": 0.0, "ytd": 0.0}
    if site_id:
        # tabela dnevnih suma u rasponu
        sql = db.text("""
          SELECT d.day, d.energy_kwh
          FROM site_energy_daily d
          WHERE d.site_id = :sid AND d.day BETWEEN :dfrom AND :dto
          ORDER BY d.day
        """)
        data = db.session.execute(sql, {"sid": site_id, "dfrom": d_from, "dto": d_to}).mappings().all()
        rows = [{"day": r["day"], "energy_kwh": float(r["energy_kwh"])} for r in data]

        # --- KPI: TODAY / MTD / YTD (računamo do d_to, da poštuje filter period) ---
        def sum_between(start: date, end: date) -> float:
            ssql = db.text("""
              SELECT COALESCE(SUM(energy_kwh), 0) AS s
              FROM site_energy_daily
              WHERE site_id = :sid AND day BETWEEN :start AND :end
            """)
            return float(db.session.execute(ssql, {"sid": site_id, "start": start, "end": end}).scalar() or 0.0)

        # Today = dan 'd_to'
        kpis["today"] = sum_between(d_to, d_to)

        # MTD = od prvog dana mjeseca d_to do d_to
        month_start = d_to.replace(day=1)
        kpis["mtd"] = sum_between(month_start, d_to)

        # YTD = od 1. januara godine d_to do d_to
        year_start = d_to.replace(month=1, day=1)
        kpis["ytd"] = sum_between(year_start, d_to)

    return render_template("reports/index.html",
                           sites=sites, rows=rows, site_id=site_id,
                           d_from=d_from, d_to=d_to, kpis=kpis)


@bp.route("/export.csv")
@login_required
def export_csv():
    site_id = request.args.get("site_id", type=int)
    d_from = _parse_date(request.args.get("from"), date.today() - timedelta(days=7))
    d_to   = _parse_date(request.args.get("to"),   date.today())

    if not site_id:
        flash("Odaberi site prije eksportovanja.", "error")
        return render_template("reports/index.html", sites=Site.query.order_by(Site.name).all(),
                               rows=[], site_id=None, d_from=d_from, d_to=d_to)

    site = Site.query.get_or_404(site_id)
    sql = db.text("""
      SELECT d.day, d.energy_kwh
      FROM site_energy_daily d
      WHERE d.site_id = :sid AND d.day BETWEEN :dfrom AND :dto
      ORDER BY d.day
    """)
    data = db.session.execute(sql, {"sid": site_id, "dfrom": d_from, "dto": d_to}).mappings().all()

    sio = StringIO()
    w = csv.writer(sio)
    w.writerow(["site", "day", "energy_kwh"])
    for r in data:
        w.writerow([site.name, r["day"].isoformat(), float(r["energy_kwh"])])
    mem = BytesIO(sio.getvalue().encode("utf-8"))
    fname = f"report_{site.name.replace(' ','_')}_{d_from}_{d_to}.csv"
    return send_file(mem, as_attachment=True, download_name=fname, mimetype="text/csv")

@bp.route("/export.xlsx")
@login_required
def export_xlsx():
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        flash("Missing package openpyxl (pip install openpyxl).", "error")
        return render_template("reports/index.html", sites=Site.query.order_by(Site.name).all(), rows=[])

    site_id = request.args.get("site_id", type=int)
    d_from = _parse_date(request.args.get("from"), date.today() - timedelta(days=7))
    d_to   = _parse_date(request.args.get("to"),   date.today())

    if not site_id:
        flash("Choose site before exporting.", "error")
        return render_template("reports/index.html", sites=Site.query.order_by(Site.name).all(),
                               rows=[], site_id=None, d_from=d_from, d_to=d_to)

    site = Site.query.get_or_404(site_id)
    sql = db.text("""
      SELECT d.day, d.energy_kwh
      FROM site_energy_daily d
      WHERE d.site_id = :sid AND d.day BETWEEN :dfrom AND :dto
      ORDER BY d.day
    """)
    data = db.session.execute(sql, {"sid": site_id, "dfrom": d_from, "dto": d_to}).mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Energy"
    ws.append(["Site", "Day", "Energy (kWh)"])
    for r in data:
        ws.append([site.name, r["day"].isoformat(), float(r["energy_kwh"])])

    # jednostavno auto-fit širine
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    mem = BytesIO()
    wb.save(mem); mem.seek(0)
    fname = f"report_{site.name.replace(' ','_')}_{d_from}_{d_to}.xlsx"
    return send_file(mem, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
